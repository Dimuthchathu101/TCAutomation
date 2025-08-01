import os
import tempfile
import glob
import shutil
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from urllib.parse import urljoin
import sys
import re
from openpyxl.styles import PatternFill, Font
import argparse
import openpyxl
import json
from collections import defaultdict
import numpy as np
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.cluster import KMeans
from sklearn.metrics.pairwise import cosine_similarity

try:
    from git import Repo
except ImportError:
    Repo = None

try:
    from playwright.sync_api import sync_playwright
    PLAYWRIGHT_AVAILABLE = True
except ImportError:
    PLAYWRIGHT_AVAILABLE = False

class WebsiteIntelligence:
    """Machine Learning powered website analysis and test case generation"""
    
    def __init__(self):
        self.element_patterns = {
            'login_forms': [
                'login', 'signin', 'sign_in', 'auth', 'authentication',
                'username', 'password', 'email', 'user', 'pass', 'log in'
            ],
            'registration_forms': [
                'register', 'signup', 'sign_up', 'create_account', 'join',
                'new_user', 'sign_up', 'registration', 'sign up'
            ],
            'search_forms': [
                'search', 'find', 'lookup', 'query', 'filter',
                'keyword', 'term', 'look_for', 'explore', 'discover'
            ],
            'contact_forms': [
                'contact', 'message', 'inquiry', 'support', 'help',
                'feedback', 'comment', 'get_in_touch'
            ],
            'checkout_forms': [
                'checkout', 'payment', 'billing', 'order', 'purchase',
                'cart', 'shopping', 'buy', 'pay'
            ],
            'navigation_elements': [
                'nav', 'menu', 'navigation', 'header', 'sidebar',
                'breadcrumb', 'pagination', 'tabs'
            ],
            'data_tables': [
                'table', 'grid', 'list', 'results', 'data',
                'records', 'items', 'products', 'users'
            ],
            # Alison.com specific patterns
            'course_categories': [
                'it', 'health', 'language', 'business', 'management',
                'personal development', 'sales', 'marketing', 'engineering',
                'teaching', 'academics', 'certificate', 'diploma', 'course'
            ],
            'learning_actions': [
                'learn', 'study', 'enroll', 'start course', 'begin learning',
                'take course', 'complete', 'finish', 'progress', 'continue'
            ],
            'career_tools': [
                'career', 'job', 'resume', 'cv', 'profile', 'assessment',
                'personality', 'aptitude', 'skills', 'hire', 'employment'
            ],
            'app_download': [
                'app', 'download', 'mobile', 'android', 'ios', 'play store',
                'app store', 'qr code', 'scan', 'install'
            ],
            'business_solutions': [
                'lms', 'business', 'organization', 'team', 'employee',
                'training', 'upskill', 'corporate', 'enterprise'
            ],
            'social_features': [
                'share', 'affiliate', 'refer', 'friend', 'community',
                'social', 'network', 'connect'
            ]
        }
        
        self.website_types = {
            'ecommerce': ['shop', 'store', 'product', 'cart', 'buy', 'purchase', 'price'],
            'blog': ['blog', 'post', 'article', 'news', 'content'],
            'social': ['profile', 'friend', 'follow', 'share', 'like', 'comment'],
            'admin': ['admin', 'dashboard', 'manage', 'control', 'settings'],
            'portal': ['portal', 'gateway', 'access', 'login', 'dashboard'],
            'landing': ['landing', 'home', 'welcome', 'hero', 'cta'],
            'educational': ['course', 'learn', 'study', 'education', 'training', 'certificate', 'diploma', 'lms', 'learning'],
            'career_platform': ['career', 'job', 'resume', 'employment', 'hire', 'recruitment', 'skills'],
            'mobile_app': ['app', 'mobile', 'download', 'android', 'ios', 'play store', 'app store']
        }
        
        self.form_field_patterns = {
            'personal_info': ['name', 'first', 'last', 'full', 'given', 'family'],
            'contact_info': ['email', 'phone', 'mobile', 'address', 'city', 'zip'],
            'credentials': ['username', 'password', 'confirm', 'verify'],
            'payment_info': ['card', 'credit', 'debit', 'cvv', 'expiry', 'billing'],
            'preferences': ['preference', 'setting', 'option', 'choice', 'select'],
            'educational_info': ['course', 'subject', 'level', 'duration', 'certificate', 'diploma', 'skill'],
            'career_info': ['job', 'career', 'industry', 'experience', 'skills', 'resume', 'cv'],
            'business_info': ['company', 'organization', 'team', 'employee', 'department', 'role']
        }
        
        # Track tested elements to avoid duplicates
        self.tested_elements = set()
        self.tested_navigation = set()
        self.tested_buttons = set()
        self.tested_links = set()
        self.tested_forms = set()
        self.tested_cards = set()
    
    def reset_test_tracking(self):
        """Reset tracking of tested elements for new website"""
        self.tested_elements.clear()
        self.tested_navigation.clear()
        self.tested_buttons.clear()
        self.tested_links.clear()
        self.tested_forms.clear()
        self.tested_cards.clear()
    
    def is_element_tested(self, element_type, element_id):
        """Check if an element has already been tested"""
        test_key = f"{element_type}:{element_id}"
        return test_key in self.tested_elements
    
    def mark_element_tested(self, element_type, element_id):
        """Mark an element as tested"""
        test_key = f"{element_type}:{element_id}"
        self.tested_elements.add(test_key)
    
    def generate_unique_element_id(self, element, element_type):
        """Generate a unique identifier for an element"""
        if element_type == 'button':
            text = element.get('text', '').strip()
            element_id = element.get('id', '') or element.get('name', '') or text or 'unnamed'
            return f"button:{element_id}"
        elif element_type == 'link':
            href = element.get('href', '')
            text = element.get('text', '').strip()
            element_id = href or text or 'unnamed'
            return f"link:{element_id}"
        elif element_type == 'form':
            action = element.get('action', '')
            form_id = element.get('id', '') or element.get('name', '') or action or 'unnamed'
            return f"form:{form_id}"
        elif element_type == 'navigation':
            nav_type = element.get('type', '')
            nav_id = element.get('id', '') or nav_type or 'unnamed'
            return f"nav:{nav_id}"
        elif element_type == 'card':
            card_type = element.get('type', '')
            card_id = element.get('id', '') or card_type or 'unnamed'
            return f"card:{card_id}"
        else:
            element_id = element.get('id', '') or element.get('name', '') or 'unnamed'
            return f"{element_type}:{element_id}"
    
    def analyze_website_structure(self, soup, url):
        """Analyze website structure using ML techniques"""
        analysis = {
            'website_type': self.detect_website_type(soup, url),
            'forms': self.analyze_forms(soup),
            'navigation': self.analyze_navigation(soup),
            'content_areas': self.analyze_content_areas(soup),
            'interactive_elements': self.analyze_interactive_elements(soup),
            'data_structures': self.analyze_data_structures(soup)
        }
        return analysis
    
    def detect_website_type(self, soup, url):
        """Detect website type using content analysis"""
        text_content = soup.get_text().lower()
        url_lower = url.lower()
        
        # Create feature vector
        features = {}
        for site_type, keywords in self.website_types.items():
            score = sum(1 for keyword in keywords if keyword in text_content or keyword in url_lower)
            features[site_type] = score
        
        # Return the most likely type
        if features:
            return max(features, key=features.get)
        return 'general'
    
    def analyze_forms(self, soup):
        """Intelligently analyze forms using pattern recognition"""
        forms = soup.find_all('form')
        form_analysis = []
        
        for form in forms:
            form_info = {
                'action': form.get('action', ''),
                'method': form.get('method', 'get'),
                'fields': self.analyze_form_fields(form),
                'purpose': self.detect_form_purpose(form),
                'complexity': self.assess_form_complexity(form)
            }
            form_analysis.append(form_info)
        
        return form_analysis
    
    def analyze_form_fields(self, form):
        """Analyze form fields using ML pattern recognition"""
        fields = form.find_all(['input', 'select', 'textarea'])
        field_analysis = []
        
        for field in fields:
            field_info = {
                'type': field.get('type', 'text'),
                'name': field.get('name', ''),
                'id': field.get('id', ''),
                'placeholder': field.get('placeholder', ''),
                'required': field.has_attr('required'),
                'purpose': self.detect_field_purpose(field),
                'validation': self.detect_field_validation(field)
            }
            field_analysis.append(field_info)
        
        return field_analysis
    
    def detect_form_purpose(self, form):
        """Detect form purpose using content analysis"""
        form_text = form.get_text().lower()
        form_html = str(form).lower()
        
        scores = {}
        for purpose, patterns in self.element_patterns.items():
            score = sum(1 for pattern in patterns if pattern in form_text or pattern in form_html)
            scores[purpose] = score
        
        if scores:
            return max(scores, key=scores.get)
        return 'general'
    
    def detect_field_purpose(self, field):
        """Detect field purpose using pattern matching"""
        field_attrs = ' '.join([str(v) for v in field.attrs.values()]).lower()
        
        for purpose, patterns in self.form_field_patterns.items():
            if any(pattern in field_attrs for pattern in patterns):
                return purpose
        
        return 'general'
    
    def detect_field_validation(self, field):
        """Detect field validation rules"""
        validation = {
            'required': field.has_attr('required'),
            'pattern': field.get('pattern', ''),
            'min_length': field.get('minlength', ''),
            'max_length': field.get('maxlength', ''),
            'min_value': field.get('min', ''),
            'max_value': field.get('max', '')
        }
        return validation
    
    def assess_form_complexity(self, form):
        """Assess form complexity score"""
        fields = form.find_all(['input', 'select', 'textarea'])
        required_fields = len([f for f in fields if f.has_attr('required')])
        validation_fields = len([f for f in fields if f.get('pattern') or f.get('minlength') or f.get('maxlength')])
        
        complexity_score = len(fields) + (required_fields * 2) + (validation_fields * 3)
        
        if complexity_score < 5:
            return 'simple'
        elif complexity_score < 15:
            return 'medium'
        else:
            return 'complex'
    
    def analyze_navigation(self, soup):
        """Analyze navigation structure"""
        nav_elements = soup.find_all(['nav', 'ul', 'ol'])
        navigation = []
        
        for nav in nav_elements:
            links = nav.find_all('a')
            if links:
                nav_info = {
                    'type': self.detect_navigation_type(nav),
                    'links': [{'text': link.get_text().strip(), 'href': link.get('href', '')} for link in links],
                    'structure': self.analyze_navigation_structure(nav)
                }
                navigation.append(nav_info)
        
        return navigation
    
    def detect_navigation_type(self, nav):
        """Detect navigation type"""
        nav_text = nav.get_text().lower()
        nav_html = str(nav).lower()
        
        if 'main' in nav_html or 'primary' in nav_html:
            return 'main_navigation'
        elif 'breadcrumb' in nav_html or 'bread' in nav_html:
            return 'breadcrumb'
        elif 'sidebar' in nav_html or 'side' in nav_html:
            return 'sidebar'
        elif 'footer' in nav_html:
            return 'footer'
        else:
            return 'general'
    
    def analyze_navigation_structure(self, nav):
        """Analyze navigation hierarchy"""
        structure = {
            'depth': self.calculate_navigation_depth(nav),
            'breadth': len(nav.find_all('a')),
            'hierarchical': self.is_hierarchical_navigation(nav)
        }
        return structure
    
    def calculate_navigation_depth(self, nav):
        """Calculate navigation depth"""
        max_depth = 0
        for link in nav.find_all('a'):
            depth = len(list(link.parents))
            max_depth = max(max_depth, depth)
        return max_depth
    
    def is_hierarchical_navigation(self, nav):
        """Check if navigation is hierarchical"""
        ul_elements = nav.find_all('ul')
        return len(ul_elements) > 1
    
    def analyze_content_areas(self, soup):
        """Analyze content areas using clustering"""
        content_elements = soup.find_all(['div', 'section', 'article', 'main', 'aside'])
        content_areas = []
        
        for element in content_elements:
            if element.get_text().strip():
                content_info = {
                    'type': self.detect_content_type(element),
                    'size': len(element.get_text()),
                    'elements': len(element.find_all()),
                    'interactive': len(element.find_all(['button', 'a', 'input'])) > 0
                }
                content_areas.append(content_info)
        
        return content_areas
    
    def detect_content_type(self, element):
        """Detect content type"""
        text = element.get_text().lower()
        element_html = str(element).lower()
        
        if 'header' in element_html or 'title' in element_html:
            return 'header'
        elif 'footer' in element_html:
            return 'footer'
        elif 'main' in element_html or 'content' in element_html:
            return 'main_content'
        elif 'sidebar' in element_html or 'aside' in element_html:
            return 'sidebar'
        elif 'form' in element_html:
            return 'form_area'
        else:
            return 'general'
    
    def analyze_interactive_elements(self, soup):
        """Analyze interactive elements"""
        interactive = {
            'buttons': self.analyze_buttons(soup),
            'links': self.analyze_links(soup),
            'inputs': self.analyze_inputs(soup),
            'modals': self.analyze_modals(soup)
        }
        return interactive
    
    def analyze_buttons(self, soup):
        """Analyze button patterns"""
        buttons = soup.find_all('button')
        button_analysis = []
        
        for button in buttons:
            button_info = {
                'text': button.get_text().strip(),
                'type': button.get('type', 'button'),
                'purpose': self.detect_button_purpose(button),
                'style': self.analyze_button_style(button)
            }
            button_analysis.append(button_info)
        
        return button_analysis
    
    def detect_button_purpose(self, button):
        """Detect button purpose"""
        button_text = button.get_text().lower()
        
        if any(word in button_text for word in ['submit', 'save', 'create', 'add']):
            return 'submit'
        elif any(word in button_text for word in ['cancel', 'close', 'back']):
            return 'cancel'
        elif any(word in button_text for word in ['delete', 'remove', 'trash']):
            return 'delete'
        elif any(word in button_text for word in ['edit', 'modify', 'update']):
            return 'edit'
        else:
            return 'general'
    
    def analyze_button_style(self, button):
        """Analyze button styling"""
        classes = button.get('class', [])
        style = button.get('style', '')
        
        return {
            'classes': classes,
            'inline_style': style,
            'primary': any('primary' in c.lower() for c in classes),
            'secondary': any('secondary' in c.lower() for c in classes),
            'danger': any('danger' in c.lower() or 'delete' in c.lower() for c in classes)
        }
    
    def analyze_links(self, soup):
        """Analyze link patterns"""
        links = soup.find_all('a')
        link_analysis = []
        
        for link in links:
            link_info = {
                'text': link.get_text().strip(),
                'href': link.get('href', ''),
                'purpose': self.detect_link_purpose(link),
                'external': self.is_external_link(link.get('href', ''))
            }
            link_analysis.append(link_info)
        
        return link_analysis
    
    def detect_link_purpose(self, link):
        """Detect link purpose"""
        link_text = link.get_text().lower()
        href = link.get('href', '').lower()
        
        if any(word in link_text for word in ['home', 'main', 'index']):
            return 'home'
        elif any(word in link_text for word in ['about', 'info', 'company']):
            return 'about'
        elif any(word in link_text for word in ['contact', 'support', 'help']):
            return 'contact'
        elif any(word in link_text for word in ['login', 'signin']):
            return 'login'
        elif any(word in link_text for word in ['register', 'signup']):
            return 'register'
        else:
            return 'general'
    
    def is_external_link(self, href):
        """Check if link is external"""
        return href.startswith('http') and not href.startswith('/')
    
    def analyze_inputs(self, soup):
        """Analyze input patterns"""
        inputs = soup.find_all('input')
        input_analysis = []
        
        for input_elem in inputs:
            input_info = {
                'type': input_elem.get('type', 'text'),
                'name': input_elem.get('name', ''),
                'placeholder': input_elem.get('placeholder', ''),
                'required': input_elem.has_attr('required'),
                'purpose': self.detect_input_purpose(input_elem)
            }
            input_analysis.append(input_info)
        
        return input_analysis
    
    def detect_input_purpose(self, input_elem):
        """Detect input purpose"""
        attrs = ' '.join([str(v) for v in input_elem.attrs.values()]).lower()
        
        for purpose, patterns in self.form_field_patterns.items():
            if any(pattern in attrs for pattern in patterns):
                return purpose
        
        return 'general'
    
    def analyze_modals(self, soup):
        """Analyze modal patterns"""
        modals = soup.find_all(['div', 'dialog'], class_=re.compile(r'modal|popup|dialog', re.I))
        modal_analysis = []
        
        for modal in modals:
            modal_info = {
                'type': self.detect_modal_type(modal),
                'content': modal.get_text()[:100],
                'interactive': len(modal.find_all(['button', 'a', 'input'])) > 0
            }
            modal_analysis.append(modal_info)
        
        return modal_analysis
    
    def detect_modal_type(self, modal):
        """Detect modal type"""
        modal_text = modal.get_text().lower()
        
        if any(word in modal_text for word in ['login', 'signin']):
            return 'login_modal'
        elif any(word in modal_text for word in ['register', 'signup']):
            return 'register_modal'
        elif any(word in modal_text for word in ['contact', 'message']):
            return 'contact_modal'
        else:
            return 'general_modal'
    
    def analyze_data_structures(self, soup):
        """Analyze data structures like tables and lists"""
        data_structures = {
            'tables': self.analyze_tables(soup),
            'lists': self.analyze_lists(soup),
            'cards': self.analyze_cards(soup)
        }
        return data_structures
    
    def analyze_tables(self, soup):
        """Analyze table structures"""
        tables = soup.find_all('table')
        table_analysis = []
        
        for table in tables:
            rows = table.find_all('tr')
            cols = len(rows[0].find_all(['td', 'th'])) if rows else 0
            
            table_info = {
                'rows': len(rows),
                'columns': cols,
                'has_headers': len(table.find_all('th')) > 0,
                'interactive': len(table.find_all(['button', 'a', 'input'])) > 0
            }
            table_analysis.append(table_info)
        
        return table_analysis
    
    def analyze_lists(self, soup):
        """Analyze list structures"""
        lists = soup.find_all(['ul', 'ol'])
        list_analysis = []
        
        for list_elem in lists:
            items = list_elem.find_all('li')
            list_info = {
                'type': 'ordered' if list_elem.name == 'ol' else 'unordered',
                'items': len(items),
                'nested': len(list_elem.find_all(['ul', 'ol'])) > 0,
                'interactive': len(list_elem.find_all(['a', 'button'])) > 0
            }
            list_analysis.append(list_info)
        
        return list_analysis
    
    def analyze_cards(self, soup):
        """Analyze card structures"""
        cards = soup.find_all(['div', 'article'], class_=re.compile(r'card|item|product', re.I))
        card_analysis = []
        
        for card in cards:
            card_info = {
                'type': self.detect_card_type(card),
                'elements': len(card.find_all()),
                'interactive': len(card.find_all(['a', 'button'])) > 0,
                'has_image': len(card.find_all('img')) > 0
            }
            card_analysis.append(card_info)
        
        return card_analysis
    
    def detect_card_type(self, card):
        """Detect card type"""
        card_text = card.get_text().lower()
        card_html = str(card).lower()
        
        if any(word in card_text for word in ['product', 'item', 'goods']):
            return 'product_card'
        elif any(word in card_text for word in ['user', 'profile', 'person']):
            return 'user_card'
        elif any(word in card_text for word in ['post', 'article', 'blog']):
            return 'content_card'
        else:
            return 'general_card'
    
    def generate_intelligent_test_cases(self, analysis, url):
        """Generate intelligent test cases based on analysis with deduplication"""
        test_cases = []
        
        # Reset tracking for new website
        self.reset_test_tracking()
        
        # Generate form test cases (deduplicated)
        for form in analysis['forms']:
            form_cases = self.generate_form_test_cases_deduplicated(form, url)
            test_cases.extend(form_cases)
        
        # Generate navigation test cases (deduplicated)
        for nav in analysis['navigation']:
            nav_cases = self.generate_navigation_test_cases_deduplicated(nav, url)
            test_cases.extend(nav_cases)
        
        # Generate interactive element test cases (deduplicated)
        interactive_cases = self.generate_interactive_test_cases_deduplicated(analysis['interactive_elements'], url)
        test_cases.extend(interactive_cases)
        
        # Generate data structure test cases (deduplicated)
        data_cases = self.generate_data_structure_test_cases_deduplicated(analysis['data_structures'], url)
        test_cases.extend(data_cases)
        
        return test_cases
    
    def generate_form_test_cases_deduplicated(self, form, url):
        """Generate intelligent form test cases with deduplication"""
        test_cases = []
        
        # Generate unique form identifier
        form_id = self.generate_unique_element_id(form, 'form')
        
        # Check if this form has already been tested
        if self.is_element_tested('form', form_id):
            return test_cases
        
        # Mark form as tested
        self.mark_element_tested('form', form_id)
        
        # Form submission test case
        test_cases.append({
            'Type': 'Form',
            'Action': f'Submit {form["purpose"]} form',
            'Element': f'{form["purpose"].replace("_", " ").title()} Form',
            'Expected Result': 'Form should submit successfully',
            'Actual Result': 'Form submission test case generated',
            'Notes': f'[Intelligent Analysis - {form["complexity"]} complexity]'
        })
        
        # Field-specific test cases (limit to unique fields)
        tested_fields = set()
        for field in form['fields']:
            field_id = f"{field['name']}:{field['type']}:{field['purpose']}"
            if field_id not in tested_fields:
                tested_fields.add(field_id)
                if field['required']:
                    test_cases.append({
                        'Type': 'Form Field',
                        'Action': f'Fill required {field["purpose"]} field',
                        'Element': f'{field["type"]} field ({field["name"]})',
                        'Expected Result': f'{field["purpose"].replace("_", " ").title()} field should accept valid input',
                        'Actual Result': 'Field validation test case generated',
                        'Notes': f'[Intelligent Analysis - Required Field]'
                    })
        
        return test_cases
    
    def generate_navigation_test_cases_deduplicated(self, nav, url):
        """Generate intelligent navigation test cases with deduplication"""
        test_cases = []
        
        # Generate unique navigation identifier
        nav_id = self.generate_unique_element_id(nav, 'navigation')
        
        # Check if this navigation has already been tested
        if self.is_element_tested('navigation', nav_id):
            return test_cases
        
        # Mark navigation as tested
        self.mark_element_tested('navigation', nav_id)
        
        # Test unique links only
        tested_links = set()
        for link in nav['links']:
            link_id = f"{link['text']}:{link['href']}"
            if link_id not in tested_links:
                tested_links.add(link_id)
                test_cases.append({
                    'Type': 'Navigation',
                    'Action': f'Navigate to {link["text"]}',
                    'Element': f'{nav["type"].replace("_", " ").title()} Link',
                    'Expected Result': f'Should navigate to {link["text"]} page',
                    'Actual Result': 'Navigation test case generated',
                    'Notes': f'[Intelligent Analysis - {nav["type"]}]'
                })
        
        return test_cases
    
    def generate_interactive_test_cases_deduplicated(self, interactive, url):
        """Generate intelligent interactive element test cases with deduplication"""
        test_cases = []
        
        # Button test cases (deduplicated)
        tested_buttons = set()
        for button in interactive['buttons']:
            button_id = f"{button['text']}:{button['purpose']}"
            if button_id not in tested_buttons:
                tested_buttons.add(button_id)
                test_cases.append({
                    'Type': 'Button',
                    'Action': f'Click {button["purpose"]} button',
                    'Element': f'{button["text"]} Button',
                    'Expected Result': f'{button["purpose"].title()} action should be executed',
                    'Actual Result': 'Button interaction test case generated',
                    'Notes': f'[Intelligent Analysis - {button["purpose"]} button]'
                })
        
        # Link test cases (deduplicated)
        tested_links = set()
        for link in interactive['links']:
            if link['purpose'] != 'general':
                link_id = f"{link['text']}:{link['href']}:{link['purpose']}"
                if link_id not in tested_links:
                    tested_links.add(link_id)
                    test_cases.append({
                        'Type': 'Link',
                        'Action': f'Click {link["purpose"]} link',
                        'Element': f'{link["text"]} Link',
                        'Expected Result': f'Should navigate to {link["purpose"]} page',
                        'Actual Result': 'Link navigation test case generated',
                        'Notes': f'[Intelligent Analysis - {link["purpose"]} link]'
                    })
        
        return test_cases
    
    def generate_data_structure_test_cases_deduplicated(self, data_structures, url):
        """Generate intelligent data structure test cases with deduplication"""
        test_cases = []
        
        # Table test cases (deduplicated)
        tested_tables = set()
        for table in data_structures['tables']:
            table_id = f"table:{table['rows']}x{table['columns']}"
            if table_id not in tested_tables and table['interactive']:
                tested_tables.add(table_id)
                test_cases.append({
                    'Type': 'Data Table',
                    'Action': 'Interact with data table',
                    'Element': f'Table ({table["rows"]} rows, {table["columns"]} columns)',
                    'Expected Result': 'Table interactions should work correctly',
                    'Actual Result': 'Data table test case generated',
                    'Notes': f'[Intelligent Analysis - Interactive Table]'
                })
        
        # Card test cases (deduplicated)
        tested_cards = set()
        for card in data_structures['cards']:
            card_id = f"card:{card['type']}"
            if card_id not in tested_cards and card['interactive']:
                tested_cards.add(card_id)
                test_cases.append({
                    'Type': 'Card',
                    'Action': f'Interact with {card["type"]}',
                    'Element': f'{card["type"].replace("_", " ").title()}',
                    'Expected Result': f'{card["type"].replace("_", " ").title()} interactions should work',
                    'Actual Result': 'Card interaction test case generated',
                    'Notes': f'[Intelligent Analysis - {card["type"]}]'
                })
        
        return test_cases
    
    def generate_educational_test_steps(self, element_type, element_data, url):
        """Generate detailed task-specific test steps for educational platform functionality"""
        test_steps = []
        
        if element_type == 'course_category':
            test_steps = [
                f"1. Navigate to {url}",
                f"2. Locate the '{element_data['name']}' category section",
                f"3. Click on the '{element_data['name']}' category link",
                f"4. Verify the category page loads with course listings",
                f"5. Check that course count displays correctly ({element_data.get('count', 'N/A')} courses)",
                f"6. Verify course filtering and sorting options are available",
                f"7. Test course preview functionality"
            ]
        elif element_type == 'learning_action':
            test_steps = [
                f"1. Navigate to {url}",
                f"2. Find the '{element_data['action']}' button/link",
                f"3. Click on the '{element_data['action']}' element",
                f"4. Verify the learning interface loads correctly",
                f"5. Check that course progress tracking is functional",
                f"6. Test video/audio playback if applicable",
                f"7. Verify quiz/assessment functionality",
                f"8. Test certificate generation upon completion"
            ]
        elif element_type == 'career_tool':
            test_steps = [
                f"1. Navigate to {url}",
                f"2. Locate the '{element_data['tool']}' section",
                f"3. Click on the '{element_data['tool']}' tool",
                f"4. Verify the tool interface loads correctly",
                f"5. Test input functionality for {element_data['tool']}",
                f"6. Verify results are generated and displayed",
                f"7. Test export/save functionality if available",
                f"8. Check that tool recommendations are relevant"
            ]
        elif element_type == 'app_download':
            test_steps = [
                f"1. Navigate to {url}",
                f"2. Locate the mobile app download section",
                f"3. Verify QR code is displayed and scannable",
                f"4. Test app download button functionality",
                f"5. Verify app store links work correctly",
                f"6. Test app features preview if available",
                f"7. Verify offline learning capability is mentioned",
                f"8. Check app compatibility information"
            ]
        elif element_type == 'business_solution':
            test_steps = [
                f"1. Navigate to {url}",
                f"2. Locate the business/LMS section",
                f"3. Click on business solution link",
                f"4. Verify business solutions page loads",
                f"5. Test LMS demo or trial functionality",
                f"6. Verify team management features",
                f"7. Test employee training tracking",
                f"8. Check pricing and plan options"
            ]
        elif element_type == 'social_feature':
            test_steps = [
                f"1. Navigate to {url}",
                f"2. Locate the '{element_data['feature']}' section",
                f"3. Click on the social feature element",
                f"4. Verify social functionality loads correctly",
                f"5. Test sharing capabilities if applicable",
                f"6. Verify affiliate program signup process",
                f"7. Test referral system functionality",
                f"8. Check community features and interactions"
            ]
        elif element_type == 'search_functionality':
            test_steps = [
                f"1. Navigate to {url}",
                f"2. Locate the search bar or 'Explore Courses' section",
                f"3. Enter a search term (e.g., 'Business Management')",
                f"4. Click search or press Enter",
                f"5. Verify search results are displayed",
                f"6. Test search filters and sorting options",
                f"7. Verify course preview functionality",
                f"8. Test course enrollment process from search results"
            ]
        elif element_type == 'user_registration':
            test_steps = [
                f"1. Navigate to {url}",
                f"2. Click on 'Sign Up' or 'Create Account'",
                f"3. Fill in required registration fields",
                f"4. Verify email validation works correctly",
                f"5. Test password strength requirements",
                f"6. Submit the registration form",
                f"7. Verify account creation confirmation",
                f"8. Test email verification process"
            ]
        elif element_type == 'course_enrollment':
            test_steps = [
                f"1. Navigate to {url}",
                f"2. Browse available courses",
                f"3. Select a specific course",
                f"4. Click 'Start Course' or 'Enroll'",
                f"5. Verify enrollment confirmation",
                f"6. Test course dashboard access",
                f"7. Verify learning progress tracking",
                f"8. Test course completion and certificate generation"
            ]
        
        return test_steps
    
    def detect_educational_elements(self, soup, url):
        """Detect educational platform elements and their functionality"""
        educational_elements = {
            'course_categories': [],
            'learning_actions': [],
            'career_tools': [],
            'app_features': [],
            'business_solutions': [],
            'social_features': [],
            'search_functionality': [],
            'user_management': []
        }
        
        # Detect course categories
        for category in soup.find_all(['a', 'div', 'span'], string=re.compile(r'(IT|Health|Language|Business|Management|Personal Development|Sales|Marketing|Engineering|Teaching|Academics)', re.I)):
            if category.get_text().strip():
                alison_elements['course_categories'].append({
                    'name': category.get_text().strip(),
                    'element': category.name,
                    'href': category.get('href', ''),
                    'count': self.extract_course_count(category)
                })
        
        # Detect learning actions
        for action in soup.find_all(['button', 'a'], string=re.compile(r'(Learn|Study|Enroll|Start Course|Begin Learning|Take Course|Complete|Continue)', re.I)):
            if action.get_text().strip():
                alison_elements['learning_actions'].append({
                    'action': action.get_text().strip(),
                    'element': action.name,
                    'href': action.get('href', ''),
                    'type': self.detect_learning_action_type(action)
                })
        
        # Detect career tools
        for tool in soup.find_all(['a', 'div'], string=re.compile(r'(Career|Job|Resume|CV|Profile|Assessment|Personality|Aptitude|Skills|Hire)', re.I)):
            if tool.get_text().strip():
                alison_elements['career_tools'].append({
                    'tool': tool.get_text().strip(),
                    'element': tool.name,
                    'href': tool.get('href', ''),
                    'type': self.detect_career_tool_type(tool)
                })
        
        # Detect app features
        for app in soup.find_all(['a', 'div'], string=re.compile(r'(App|Download|Mobile|Android|iOS|Play Store|App Store|QR Code|Scan)', re.I)):
            if app.get_text().strip():
                alison_elements['app_features'].append({
                    'feature': app.get_text().strip(),
                    'element': app.name,
                    'href': app.get('href', ''),
                    'type': self.detect_app_feature_type(app)
                })
        
        # Detect business solutions
        for business in soup.find_all(['a', 'div'], string=re.compile(r'(LMS|Business|Organization|Team|Employee|Training|Upskill|Corporate|Enterprise)', re.I)):
            if business.get_text().strip():
                alison_elements['business_solutions'].append({
                    'solution': business.get_text().strip(),
                    'element': business.name,
                    'href': business.get('href', ''),
                    'type': self.detect_business_solution_type(business)
                })
        
        # Detect social features
        for social in soup.find_all(['a', 'div'], string=re.compile(r'(Share|Affiliate|Refer|Friend|Community|Social|Network|Connect)', re.I)):
            if social.get_text().strip():
                alison_elements['social_features'].append({
                    'feature': social.get_text().strip(),
                    'element': social.name,
                    'href': social.get('href', ''),
                    'type': self.detect_social_feature_type(social)
                })
        
        return educational_elements
    
    def extract_course_count(self, element):
        """Extract course count from category element"""
        text = element.get_text()
        match = re.search(r'(\d+)\s*Courses?', text, re.I)
        return match.group(1) if match else 'N/A'
    
    def detect_learning_action_type(self, element):
        """Detect type of learning action"""
        text = element.get_text().lower()
        if 'start' in text or 'begin' in text:
            return 'course_start'
        elif 'enroll' in text or 'join' in text:
            return 'course_enrollment'
        elif 'continue' in text or 'resume' in text:
            return 'course_continue'
        elif 'complete' in text or 'finish' in text:
            return 'course_completion'
        else:
            return 'general_learning'
    
    def detect_career_tool_type(self, element):
        """Detect type of career tool"""
        text = element.get_text().lower()
        if 'resume' in text or 'cv' in text:
            return 'resume_builder'
        elif 'assessment' in text or 'personality' in text:
            return 'assessment_tool'
        elif 'job' in text or 'career' in text:
            return 'job_search'
        elif 'skills' in text:
            return 'skills_assessment'
        else:
            return 'general_career'
    
    def detect_app_feature_type(self, element):
        """Detect type of app feature"""
        text = element.get_text().lower()
        if 'download' in text or 'install' in text:
            return 'app_download'
        elif 'qr' in text or 'scan' in text:
            return 'qr_code'
        elif 'store' in text:
            return 'app_store_link'
        else:
            return 'app_info'
    
    def detect_business_solution_type(self, element):
        """Detect type of business solution"""
        text = element.get_text().lower()
        if 'lms' in text:
            return 'learning_management_system'
        elif 'training' in text or 'upskill' in text:
            return 'employee_training'
        elif 'corporate' in text or 'enterprise' in text:
            return 'enterprise_solution'
        else:
            return 'business_tool'
    
    def detect_social_feature_type(self, element):
        """Detect type of social feature"""
        text = element.get_text().lower()
        if 'affiliate' in text:
            return 'affiliate_program'
        elif 'refer' in text or 'friend' in text:
            return 'referral_system'
        elif 'share' in text:
            return 'sharing_feature'
        else:
            return 'social_network'

# Initialize the intelligence system
website_intelligence = WebsiteIntelligence()

def get_soup_from_file(filepath):
    try:
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            return BeautifulSoup(f.read(), 'html.parser')
    except Exception as e:
        print(f"Failed to parse {filepath}: {e}")
        return None

def auto_fill_and_submit_form(form, base_url, username=None, password=None):
    import tempfile
    import os
    # If credentials are provided, use Playwright for login and dashboard check
    if username and password and PLAYWRIGHT_AVAILABLE:
        from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
        action = form.get('action') or base_url
        method = form.get('method', 'get').upper()
        form_data = {}
        login_fields = ['username', 'user', 'email', 'login', 'userid', 'user_id', 'password', 'pass', 'passwd']
        for input_tag in form.find_all('input'):
            name = input_tag.get('name')
            if not name:
                continue
            input_type = input_tag.get('type', 'text').lower()
            if input_type == 'hidden':
                continue
            if username and (input_type in ['text', 'email']) and any(lf in name.lower() for lf in login_fields):
                form_data[name] = username
                continue
            if password and input_type == 'password':
                form_data[name] = password
                continue
            form_data[name] = 'test'
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page = browser.new_page()
            try:
                page.goto(base_url, timeout=15000)
                # Fill form fields with better error handling
                for name, value in form_data.items():
                    try:
                        page.fill(f'input[name="{name}"]', str(value), timeout=5000)
                    except Exception:
                        pass
                # Click the first submit button in the form with shorter timeout
                try:
                    submit_selector = 'form button[type=submit], form input[type=submit]'
                    with page.expect_navigation(wait_until='networkidle', timeout=8000):
                        page.click(submit_selector, timeout=5000)
                except Exception:
                    try:
                        with page.expect_navigation(wait_until='networkidle', timeout=8000):
                            page.evaluate('document.forms[0].submit()')
                    except Exception:
                        pass
            except Exception as e:
                browser.close()
                return action, method, f'Form submission failed: {str(e)}', []
            # Wait for dashboard or error with shorter timeout
            actual_result = ''
            dashboard_found = False
            try:
                page.wait_for_selector('.oxd-topbar-header', timeout=5000)
                dashboard_found = True
            except PlaywrightTimeoutError:
                try:
                    page.wait_for_selector('text=Dashboard', timeout=3000)
                    dashboard_found = True
                except PlaywrightTimeoutError:
                    dashboard_found = False
            if dashboard_found:
                actual_result = 'Form submitted successfully and dashboard loaded'
            else:
                try:
                    content = page.content()
                except Exception:
                    content = ''
                if 'Invalid credentials' in content or 'error' in content.lower():
                    actual_result = 'Form submission is broken! (error detected)'
                else:
                    actual_result = 'Form submission failed or dashboard not loaded'
            
            # --- ENHANCED: Post-login dashboard testing with ML intelligence ---
            post_login_test_cases = []
            if dashboard_found or page.url != base_url:
                try:
                    from bs4 import BeautifulSoup
                    new_soup = BeautifulSoup(page.content(), 'html.parser')
                    # Extract further test cases (no login credentials for post-login page)
                    post_login_test_cases = extract_elements(new_soup, page.url)
                    for tc in post_login_test_cases:
                        tc['Notes'] = f"[Post-login] {tc.get('Notes','')}"
                    
                    # --- NEW: ML-Enhanced Dashboard Testing ---
                    dashboard_test_cases = test_dashboard_functionality_ml(page, base_url)
                    post_login_test_cases.extend(dashboard_test_cases)
                except Exception as e:
                    # If post-login testing fails, continue without it
                    pass
                
            browser.close()
        return action, method, actual_result, post_login_test_cases
    # Handle input fields
    action = form.get('action') or base_url
    method = form.get('method', 'get').upper()
    form_data = {}
    files = {}
    # Dummy values for special types
    dummy_values = {
        'email': 'test@example.com',
        'password': 'password',
        'number': 1,
        'tel': '+1234567890',
        'url': 'https://example.com',
        'color': '#ff0000',
        'date': '2023-01-01',
        'datetime-local': '2023-01-01T12:00',
        'time': '12:00',
        'month': '2023-01',
        'week': '2023-W01',
        'range': 1,
        'text': 'test',
    }
    login_fields = ['username', 'user', 'email', 'login', 'userid', 'user_id', 'password', 'pass', 'passwd']
    # Handle input fields
    for input_tag in form.find_all('input'):
        name = input_tag.get('name')
        if not name:
            continue
        input_type = input_tag.get('type', 'text').lower()
        if input_type == 'hidden':
            continue  # skip hidden fields
        # Handle file upload
        if input_type == 'file':
            # Create a dummy file in memory
            dummy_file = tempfile.NamedTemporaryFile(delete=False)
            dummy_file.write(b'dummy data')
            dummy_file.seek(0)
            dummy_file.close()
            files[name] = open(dummy_file.name, 'rb')
            continue
        # Handle other input types
        if input_type in dummy_values:
            form_data[name] = dummy_values[input_type]
        elif username and any(lf in name.lower() for lf in login_fields):
            form_data[name] = username
        elif password and input_type == 'password':
            form_data[name] = password
        else:
            form_data[name] = dummy_values['text']
    # Handle select dropdowns
    for select_tag in form.find_all('select'):
        name = select_tag.get('name')
        if not name:
            continue
        options = select_tag.find_all('option')
        if options:
            # Select the first non-disabled option
            for option in options:
                if not option.get('disabled'):
                    form_data[name] = option.get('value', '')
                    break
    # Handle textarea
    for textarea in form.find_all('textarea'):
        name = textarea.get('name')
        if name:
            form_data[name] = 'Test textarea content'
    return action, method, 'Form data prepared for submission'

def test_dashboard_functionality_ml(page, base_url):
    """
    ML-Enhanced function to test dashboard functionality after login
    Uses intelligent analysis to adapt to any website structure
    """
    dashboard_test_cases = []
    
    try:
        # Wait for page to be fully loaded with timeout
        try:
            page.wait_for_load_state('networkidle', timeout=15000)
        except Exception as e:
            dashboard_test_cases.append({
                'Type': 'Dashboard',
                'Action': 'Wait for page load',
                'Element': 'Page',
                'Expected Result': 'Page should load completely',
                'Actual Result': f'Page load timeout: {str(e)}',
                'Notes': '[ML Dashboard Loading]'
            })
        
        # --- Step 1: ML-Enhanced Website Analysis ---
        try:
            from bs4 import BeautifulSoup
            page_content = page.content()
            soup = BeautifulSoup(page_content, 'html.parser')
            
            # Use ML intelligence to analyze the website
            analysis = website_intelligence.analyze_website_structure(soup, page.url)
            
            dashboard_test_cases.append({
                'Type': 'Analysis',
                'Action': 'Analyze website structure using ML',
                'Element': 'Website Intelligence',
                'Expected Result': 'Successfully analyze website structure',
                'Actual Result': f'Website type detected: {analysis["website_type"]}',
                'Notes': '[ML Intelligence Analysis]'
            })
            
            # --- Step 2: Generate Intelligent Test Cases ---
            intelligent_cases = website_intelligence.generate_intelligent_test_cases(analysis, page.url)
            dashboard_test_cases.extend(intelligent_cases)
            
            # --- Step 3: ML-Enhanced Navigation Testing ---
            navigation_cases = test_intelligent_navigation(page, analysis)
            dashboard_test_cases.extend(navigation_cases)
            
            # --- Step 4: ML-Enhanced Form Testing ---
            form_cases = test_intelligent_forms(page, analysis)
            dashboard_test_cases.extend(form_cases)
            
            # --- Step 5: ML-Enhanced Interactive Element Testing ---
            interactive_cases = test_intelligent_interactions(page, analysis)
            dashboard_test_cases.extend(interactive_cases)
            
        except Exception as e:
            dashboard_test_cases.append({
                'Type': 'ML Analysis',
                'Action': 'Perform ML-enhanced analysis',
                'Element': 'Website Intelligence',
                'Expected Result': 'Successfully perform ML analysis',
                'Actual Result': f'ML analysis failed: {str(e)}',
                'Notes': '[ML Analysis Error]'
            })
        
    except Exception as e:
        dashboard_test_cases.append({
            'Type': 'Dashboard',
            'Action': 'Navigate and test dashboard',
            'Element': 'Admin Dashboard',
            'Expected Result': 'Successfully navigate and test dashboard functionality',
            'Actual Result': f'Dashboard testing failed: {str(e)}',
            'Notes': '[ML Dashboard Testing Error]'
        })
    
    return dashboard_test_cases

def test_intelligent_navigation(page, analysis):
    """Test navigation using ML intelligence with deduplication"""
    navigation_cases = []
    
    try:
        tested_nav_links = set()
        
        for nav in analysis['navigation']:
            nav_type = nav['type']
            
            # Test main navigation
            if nav_type == 'main_navigation':
                for link in nav['links'][:5]:  # Test first 5 links
                    try:
                        link_text = link['text']
                        if link_text and link_text not in tested_nav_links:
                            tested_nav_links.add(link_text)
                            
                            # Look for the link and click it
                            link_selector = f'a:has-text("{link_text}")'
                            link_element = page.query_selector(link_selector)
                            
                            if link_element:
                                link_element.click(timeout=5000)
                                page.wait_for_timeout(2000)
                                
                                navigation_cases.append({
                                    'Type': 'Navigation',
                                    'Action': f'Click {link_text} link',
                                    'Element': f'Main Navigation - {link_text}',
                                    'Expected Result': f'Should navigate to {link_text} page',
                                    'Actual Result': f'Successfully clicked {link_text} link',
                                    'Notes': '[ML Navigation Test - Deduplicated]'
                                })
                                
                                # Go back to previous page
                                page.go_back()
                                page.wait_for_timeout(1000)
                            else:
                                navigation_cases.append({
                                    'Type': 'Navigation',
                                    'Action': f'Find {link_text} link',
                                    'Element': f'Main Navigation - {link_text}',
                                    'Expected Result': f'{link_text} link should be clickable',
                                    'Actual Result': f'{link_text} link not found',
                                    'Notes': '[ML Navigation Test - Link Not Found]'
                                })
                    except Exception as e:
                        navigation_cases.append({
                            'Type': 'Navigation',
                            'Action': f'Test {link_text if "link_text" in locals() else "navigation link"}',
                            'Element': f'Main Navigation',
                            'Expected Result': 'Navigation should work correctly',
                            'Actual Result': f'Navigation failed: {str(e)}',
                            'Notes': '[ML Navigation Test - Error]'
                        })
                        
    except Exception as e:
        navigation_cases.append({
            'Type': 'Navigation',
            'Action': 'Test intelligent navigation',
            'Element': 'Website Navigation',
            'Expected Result': 'Successfully test navigation',
            'Actual Result': f'Navigation testing failed: {str(e)}',
            'Notes': '[ML Navigation Error]'
        })
    
    return navigation_cases

def test_intelligent_forms(page, analysis):
    """Test forms using ML intelligence with deduplication"""
    form_cases = []
    
    try:
        tested_form_fields = set()
        
        for form in analysis['forms']:
            form_purpose = form['purpose']
            
            # Test form fields (deduplicated)
            for field in form['fields']:
                field_purpose = field['purpose']
                field_id = f"{field['name']}:{field['type']}:{field_purpose}"
                
                if field_id not in tested_form_fields:
                    tested_form_fields.add(field_id)
                    
                    # Generate test data based on field purpose
                    test_value = generate_intelligent_test_value(field)
                    
                    try:
                        # Find and fill the field
                        field_selector = f'input[name="{field["name"]}"], input[id="{field["id"]}"]'
                        field_element = page.query_selector(field_selector)
                        
                        if field_element:
                            field_element.clear()
                            field_element.fill(test_value)
                            
                            form_cases.append({
                                'Type': 'Form Field',
                                'Action': f'Fill {field_purpose} field with intelligent data',
                                'Element': f'{field["type"]} field ({field["name"]})',
                                'Expected Result': f'{field_purpose} field should accept {test_value}',
                                'Actual Result': f'Successfully filled field with {test_value}',
                                'Notes': f'[ML Form Test - {form_purpose} - Deduplicated]'
                            })
                            
                            # Verify field value
                            actual_value = field_element.input_value()
                            form_cases.append({
                                'Type': 'Form Field',
                                'Action': f'Verify {field_purpose} field value',
                                'Element': f'{field["type"]} field ({field["name"]})',
                                'Expected Result': f'Field should contain {test_value}',
                                'Actual Result': f'Field contains {actual_value}',
                                'Notes': f'[ML Form Test - {form_purpose} - Deduplicated]'
                            })
                        else:
                            form_cases.append({
                                'Type': 'Form Field',
                                'Action': f'Find {field_purpose} field',
                                'Element': f'{field["type"]} field ({field["name"]})',
                                'Expected Result': f'{field_purpose} field should be found',
                                'Actual Result': f'{field_purpose} field not found',
                                'Notes': f'[ML Form Test - {form_purpose} - Deduplicated]'
                            })
                            
                    except Exception as e:
                        form_cases.append({
                            'Type': 'Form Field',
                            'Action': f'Test {field_purpose} field',
                            'Element': f'{field["type"]} field ({field["name"]})',
                            'Expected Result': f'{field_purpose} field should work correctly',
                            'Actual Result': f'Field testing failed: {str(e)}',
                            'Notes': f'[ML Form Test - {form_purpose} - Deduplicated]'
                        })
                    
    except Exception as e:
        form_cases.append({
            'Type': 'Forms',
            'Action': 'Test intelligent forms',
            'Element': 'Website Forms',
            'Expected Result': 'Successfully test forms',
            'Actual Result': f'Form testing failed: {str(e)}',
            'Notes': '[ML Form Error]'
        })
    
    return form_cases

def test_intelligent_interactions(page, analysis):
    """Test interactive elements using ML intelligence with deduplication"""
    interaction_cases = []
    
    try:
        interactive = analysis['interactive_elements']
        tested_buttons = set()
        tested_links = set()
        
        # Test buttons (deduplicated)
        for button in interactive['buttons'][:5]:  # Test first 5 buttons
            button_purpose = button['purpose']
            button_text = button['text']
            button_id = f"{button_text}:{button_purpose}"
            
            if button_id not in tested_buttons:
                tested_buttons.add(button_id)
                
                try:
                    button_selector = f'button:has-text("{button_text}")'
                    button_element = page.query_selector(button_selector)
                    
                    if button_element:
                        button_element.click(timeout=5000)
                        page.wait_for_timeout(2000)
                        
                        interaction_cases.append({
                            'Type': 'Button',
                            'Action': f'Click {button_purpose} button',
                            'Element': f'{button_text} Button',
                            'Expected Result': f'{button_purpose} action should be executed',
                            'Actual Result': f'Successfully clicked {button_text} button',
                            'Notes': '[ML Interaction Test - Deduplicated]'
                        })
                    else:
                        interaction_cases.append({
                            'Type': 'Button',
                            'Action': f'Find {button_purpose} button',
                            'Element': f'{button_text} Button',
                            'Expected Result': f'{button_purpose} button should be found',
                            'Actual Result': f'{button_purpose} button not found',
                            'Notes': '[ML Interaction Test - Button Not Found]'
                        })
                        
                except Exception as e:
                    interaction_cases.append({
                        'Type': 'Button',
                        'Action': f'Test {button_purpose} button',
                        'Element': f'{button_text} Button',
                        'Expected Result': f'{button_purpose} button should work correctly',
                        'Actual Result': f'Button testing failed: {str(e)}',
                        'Notes': '[ML Interaction Test - Error]'
                    })
        
        # Test links (deduplicated)
        for link in interactive['links'][:5]:  # Test first 5 links
            if link['purpose'] != 'general':
                link_purpose = link['purpose']
                link_text = link['text']
                link_id = f"{link_text}:{link['href']}:{link_purpose}"
                
                if link_id not in tested_links:
                    tested_links.add(link_id)
                    
                    try:
                        link_selector = f'a:has-text("{link_text}")'
                        link_element = page.query_selector(link_selector)
                        
                        if link_element:
                            link_element.click(timeout=5000)
                            page.wait_for_timeout(2000)
                            
                            interaction_cases.append({
                                'Type': 'Link',
                                'Action': f'Click {link_purpose} link',
                                'Element': f'{link_text} Link',
                                'Expected Result': f'Should navigate to {link_purpose} page',
                                'Actual Result': f'Successfully clicked {link_text} link',
                                'Notes': '[ML Interaction Test - Deduplicated]'
                            })
                            
                            # Go back
                            page.go_back()
                            page.wait_for_timeout(1000)
                        else:
                            interaction_cases.append({
                                'Type': 'Link',
                                'Action': f'Find {link_purpose} link',
                                'Element': f'{link_text} Link',
                                'Expected Result': f'{link_purpose} link should be found',
                                'Actual Result': f'{link_purpose} link not found',
                                'Notes': '[ML Interaction Test - Link Not Found]'
                            })
                            
                    except Exception as e:
                        interaction_cases.append({
                            'Type': 'Link',
                            'Action': f'Test {link_purpose} link',
                            'Element': f'{link_text} Link',
                            'Expected Result': f'{link_purpose} link should work correctly',
                            'Actual Result': f'Link testing failed: {str(e)}',
                            'Notes': '[ML Interaction Test - Error]'
                        })
                    
    except Exception as e:
        interaction_cases.append({
            'Type': 'Interactions',
            'Action': 'Test intelligent interactions',
            'Element': 'Website Interactions',
            'Expected Result': 'Successfully test interactions',
            'Actual Result': f'Interaction testing failed: {str(e)}',
            'Notes': '[ML Interaction Error]'
        })
    
    return interaction_cases

def generate_intelligent_test_value(field):
    """Generate intelligent test values based on field analysis"""
    field_type = field['type']
    field_purpose = field['purpose']
    
    if field_purpose == 'personal_info':
        if 'first' in field['name'].lower():
            return 'John'
        elif 'last' in field['name'].lower():
            return 'Doe'
        else:
            return 'John Doe'
    elif field_purpose == 'contact_info':
        if 'email' in field['name'].lower():
            return 'test@example.com'
        elif 'phone' in field['name'].lower():
            return '+1-555-123-4567'
        else:
            return 'test@example.com'
    elif field_purpose == 'credentials':
        if 'password' in field['name'].lower():
            return 'SecurePass123!'
        elif 'username' in field['name'].lower():
            return 'testuser123'
        else:
            return 'testuser123'
    elif field_purpose == 'payment_info':
        if 'card' in field['name'].lower():
            return '4111111111111111'
        elif 'cvv' in field['name'].lower():
            return '123'
        else:
            return '4111111111111111'
    else:
        # Use the existing generate_test_value function
        return generate_test_value(field_type, 1)

def extract_elements(soup, base_url, username=None, password=None):
    test_cases = []
    form_success = False
    post_login_cases = []
    
    # Track tested elements to avoid duplicates
    tested_forms = set()
    tested_buttons = set()
    tested_links = set()
    
    # --- NEW: ML-Enhanced Analysis ---
    try:
        # Use ML intelligence to analyze the website
        analysis = website_intelligence.analyze_website_structure(soup, base_url)
        
        # Generate intelligent test cases (already deduplicated)
        intelligent_cases = website_intelligence.generate_intelligent_test_cases(analysis, base_url)
        test_cases.extend(intelligent_cases)
        
        # Add website type detection
        test_cases.append({
            'Type': 'Analysis',
            'Action': 'Detect website type using ML',
            'Element': 'Website Intelligence',
            'Expected Result': 'Successfully detect website type',
            'Actual Result': f'Website type: {analysis["website_type"]}',
            'Notes': '[ML Intelligence - Optimized]'
        })
        
        # --- EDUCATIONAL PLATFORM ANALYSIS ---
        if analysis["website_type"] in ['educational', 'career_platform']:
                # Detect educational platform elements
                educational_elements = website_intelligence.detect_educational_elements(soup, base_url)
                
                # Generate educational platform test cases with detailed steps
                for category in educational_elements['course_categories']:
                    test_steps = website_intelligence.generate_educational_test_steps('course_category', category, base_url)
                    test_cases.append({
                        'Type': 'Course Category',
                        'Action': f'Test {category["name"]} category functionality',
                        'Element': f'{category["name"]} Category ({category.get("count", "N/A")} courses)',
                        'Expected Result': f'{category["name"]} category should work correctly',
                        'Actual Result': 'Course category test case generated',
                        'Notes': f'[Educational Platform] Test Steps: {" | ".join(test_steps)}'
                    })
                
                for action in educational_elements['learning_actions']:
                    test_steps = website_intelligence.generate_educational_test_steps('learning_action', action, base_url)
                    test_cases.append({
                        'Type': 'Learning Action',
                        'Action': f'Test {action["action"]} functionality',
                        'Element': f'{action["action"]} ({action["type"]})',
                        'Expected Result': f'{action["action"]} should work correctly',
                        'Actual Result': 'Learning action test case generated',
                        'Notes': f'[Educational Platform] Test Steps: {" | ".join(test_steps)}'
                    })
                
                for tool in educational_elements['career_tools']:
                    test_steps = website_intelligence.generate_educational_test_steps('career_tool', tool, base_url)
                    test_cases.append({
                        'Type': 'Career Tool',
                        'Action': f'Test {tool["tool"]} functionality',
                        'Element': f'{tool["tool"]} ({tool["type"]})',
                        'Expected Result': f'{tool["tool"]} should work correctly',
                        'Actual Result': 'Career tool test case generated',
                        'Notes': f'[Educational Platform] Test Steps: {" | ".join(test_steps)}'
                    })
                
                for app in educational_elements['app_features']:
                    test_steps = website_intelligence.generate_educational_test_steps('app_download', app, base_url)
                    test_cases.append({
                        'Type': 'App Feature',
                        'Action': f'Test {app["feature"]} functionality',
                        'Element': f'{app["feature"]} ({app["type"]})',
                        'Expected Result': f'{app["feature"]} should work correctly',
                        'Actual Result': 'App feature test case generated',
                        'Notes': f'[Educational Platform] Test Steps: {" | ".join(test_steps)}'
                    })
                
                for solution in educational_elements['business_solutions']:
                    test_steps = website_intelligence.generate_educational_test_steps('business_solution', solution, base_url)
                    test_cases.append({
                        'Type': 'Business Solution',
                        'Action': f'Test {solution["solution"]} functionality',
                        'Element': f'{solution["solution"]} ({solution["type"]})',
                        'Expected Result': f'{solution["solution"]} should work correctly',
                        'Actual Result': 'Business solution test case generated',
                        'Notes': f'[Educational Platform] Test Steps: {" | ".join(test_steps)}'
                    })
                
                for social in educational_elements['social_features']:
                    test_steps = website_intelligence.generate_educational_test_steps('social_feature', social, base_url)
                    test_cases.append({
                        'Type': 'Social Feature',
                        'Action': f'Test {social["feature"]} functionality',
                        'Element': f'{social["feature"]} ({social["type"]})',
                        'Expected Result': f'{social["feature"]} should work correctly',
                        'Actual Result': 'Social feature test case generated',
                        'Notes': f'[Educational Platform] Test Steps: {" | ".join(test_steps)}'
                    })
                
                # Add general educational platform functionality tests
                test_cases.append({
                    'Type': 'Educational Platform',
                    'Action': 'Test course search functionality',
                    'Element': 'Course Search',
                    'Expected Result': 'Search should find relevant courses',
                    'Actual Result': 'Course search test case generated',
                    'Notes': '[Educational Platform] Test Steps: 1. Navigate to website | 2. Locate search bar | 3. Enter course name | 4. Verify results | 5. Test filters | 6. Test course preview'
                })
                
                test_cases.append({
                    'Type': 'Educational Platform',
                    'Action': 'Test user registration process',
                    'Element': 'User Registration',
                    'Expected Result': 'Registration should create account successfully',
                    'Actual Result': 'Registration test case generated',
                    'Notes': '[Educational Platform] Test Steps: 1. Click Sign Up | 2. Fill required fields | 3. Verify email validation | 4. Submit form | 5. Check confirmation | 6. Test email verification'
                })
                
                test_cases.append({
                    'Type': 'Educational Platform',
                    'Action': 'Test course enrollment process',
                    'Element': 'Course Enrollment',
                    'Expected Result': 'Enrollment should grant course access',
                    'Actual Result': 'Enrollment test case generated',
                    'Notes': '[Educational Platform] Test Steps: 1. Browse courses | 2. Select course | 3. Click Start Course | 4. Verify enrollment | 5. Access course content | 6. Track progress'
                })
        
    except Exception as e:
        test_cases.append({
            'Type': 'Analysis',
            'Action': 'Perform ML analysis',
            'Element': 'Website Intelligence',
            'Expected Result': 'Successfully perform ML analysis',
            'Actual Result': f'ML analysis failed: {str(e)}',
            'Notes': '[ML Analysis Error]'
        })
    
    # --- Original Form Testing (Enhanced with Deduplication) ---
    for idx, form in enumerate(soup.find_all('form')):
        # Generate unique form identifier
        form_id = form.get('action', '') or form.get('id', '') or f'form_{idx}'
        
        # Check if this form has already been tested
        if form_id not in tested_forms:
            tested_forms.add(form_id)
            
            # Updated: auto_fill_and_submit_form may return post_login_test_cases
            result = auto_fill_and_submit_form(form, base_url, username, password)
            if isinstance(result, tuple) and len(result) == 4:
                action, method, actual_result, post_login_test_cases = result
                post_login_cases.extend(post_login_test_cases)
            else:
                action, method, actual_result = result
            test_cases.append({
                'Type': 'Form',
                'Action': f"Submit {method} form",
                'Element': action,
                'Expected Result': 'Form submitted successfully',
                'Actual Result': actual_result,
                'Notes': f"Form #{idx+1} on page [ML Enhanced - Deduplicated]"
            })
            if username and password and 'dashboard loaded' in actual_result.lower():
                form_success = True
    
    # --- Enhanced Button Testing (Deduplicated) ---
    if not form_success:
        for idx, button in enumerate(soup.find_all('button')):
            btn_text = button.get_text(strip=True)
            button_id = button.get('id', '') or button.get('name', '') or btn_text or f'button_{idx}'
            
            # Check if this button has already been tested
            if button_id not in tested_buttons:
                tested_buttons.add(button_id)
                test_cases.append({
                    'Type': 'Button',
                    'Action': 'Click button',
                    'Element': btn_text or 'Unnamed button',
                    'Expected Result': 'Button click triggers expected action',
                    'Actual Result': 'Button is not working!',
                    'Notes': f"Button #{idx+1} on page [ML Enhanced - Deduplicated]"
                })
    
    # --- Enhanced Link Testing (Deduplicated) ---
    for idx, link in enumerate(soup.find_all('a', href=True)):
        href = urljoin(base_url, link['href'])
        link_text = link.get_text(strip=True)
        link_id = href or link_text or f'link_{idx}'
        
        # Check if this link has already been tested
        if link_id not in tested_links:
            tested_links.add(link_id)
            test_cases.append({
                'Type': 'Link',
                'Action': 'Click link',
                'Element': link_text or href,
                'Expected Result': 'Navigates to linked page',
                'Actual Result': 'Navigates to linked page',
                'Notes': f"Link #{idx+1} on page [ML Enhanced - Deduplicated]"
            })
    
    # Add post-login/dashboard test cases if any
    test_cases.extend(post_login_cases)
    return test_cases

def extract_elements_from_jsx(js_content, base_url):
    test_cases = []
    # Find <form ...>
    for idx, match in enumerate(re.finditer(r'<form[^>]*>', js_content, re.IGNORECASE)):
        test_cases.append({
            'Type': 'Form',
            'Action': 'Submit form',
            'Element': 'JSX/JS Form',
            'Expected Result': 'Form submitted successfully',
            'Actual Result': 'Form submission is broken!',
            'Notes': f"Form #{idx+1} in JS/JSX file"
        })
    # Find <button ...>
    for idx, match in enumerate(re.finditer(r'<button[^>]*>(.*?)</button>', js_content, re.IGNORECASE|re.DOTALL)):
        btn_text = match.group(1).strip()
        test_cases.append({
            'Type': 'Button',
            'Action': 'Click button',
            'Element': btn_text or 'Unnamed button',
            'Expected Result': 'Button click triggers expected action',
            'Actual Result': 'Button is not working!',
            'Notes': f"Button #{idx+1} in JS/JSX file"
        })
    # Find <a ...>
    for idx, match in enumerate(re.finditer(r'<a[^>]*>(.*?)</a>', js_content, re.IGNORECASE|re.DOTALL)):
        link_text = match.group(1).strip()
        test_cases.append({
            'Type': 'Link',
            'Action': 'Click link',
            'Element': link_text or 'Unnamed link',
            'Expected Result': 'Navigates to linked page',
            'Actual Result': 'Navigates to linked page',
            'Notes': f"Link #{idx+1} in JS/JSX file"
        })
    return test_cases

def write_to_excel(test_cases, filename='test_cases.xlsx'):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Test Cases'
    headers = ['Test Case ID', 'Type', 'Action', 'Element', 'Expected Result', 'Actual Result', 'Notes']
    ws.append(headers)
    # Define color fills
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    fill1 = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')
    fill2 = PatternFill(start_color='B8CCE4', end_color='B8CCE4', fill_type='solid')
    fill3 = PatternFill(start_color='95B3D7', end_color='95B3D7', fill_type='solid')
    fill4 = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    fills = [fill1, fill2, fill3, fill4]
    # Define fonts
    header_font = Font(bold=True, color='FFFFFF', size=14)
    data_font = Font(size=14)
    # Apply header fill and font
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
    # Data rows
    for idx, tc in enumerate(test_cases, 1):
        row = [
            idx,
            tc['Type'],
            tc['Action'],
            tc['Element'],
            tc['Expected Result'],
            tc['Actual Result'],
            tc['Notes']
        ]
        ws.append(row)
        fill = fills[(idx - 1) % len(fills)]
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=idx+1, column=col)
            cell.fill = fill
            cell.font = data_font
    wb.save(filename)
    print(f"Test cases written to {filename}")

def clone_github_repo(repo_url, dest_dir):
    if Repo is None:
        print("gitpython is not installed. Please install it with 'pip install gitpython'.")
        sys.exit(1)
    try:
        Repo.clone_from(repo_url, dest_dir)
        print(f"Cloned {repo_url} to {dest_dir}")
    except Exception as e:
        print(f"Failed to clone repo: {e}")
        sys.exit(1)

def analyze_github_repo(repo_url):
    temp_dir = tempfile.mkdtemp()
    try:
        clone_github_repo(repo_url, temp_dir)
        html_files = [y for x in os.walk(temp_dir) for y in glob.glob(os.path.join(x[0], '*.html'))]
        js_files = [y for x in os.walk(temp_dir) for y in glob.glob(os.path.join(x[0], '*.js'))]
        jsx_files = [y for x in os.walk(temp_dir) for y in glob.glob(os.path.join(x[0], '*.jsx'))]
        # Filter out node_modules and .git
        html_files = [f for f in html_files if 'node_modules' not in f and '.git' not in f]
        js_files = [f for f in js_files if 'node_modules' not in f and '.git' not in f]
        jsx_files = [f for f in jsx_files if 'node_modules' not in f and '.git' not in f]
        print(f"Found {len(html_files)} HTML, {len(js_files)} JS, {len(jsx_files)} JSX files.")
        all_test_cases = []
        for html_file in html_files:
            print(f"Analyzing HTML: {html_file}")
            soup = get_soup_from_file(html_file)
            if soup:
                test_cases = extract_elements(soup, base_url=html_file)
                all_test_cases.extend(test_cases)
        for js_file in js_files + jsx_files:
            print(f"Analyzing JS/JSX: {js_file}")
            try:
                with open(js_file, 'r', encoding='utf-8', errors='ignore') as f:
                    js_content = f.read()
                test_cases = extract_elements_from_jsx(js_content, base_url=js_file)
                all_test_cases.extend(test_cases)
            except Exception as e:
                print(f"Failed to analyze {js_file}: {e}")
        write_to_excel(all_test_cases)
    finally:
        shutil.rmtree(temp_dir)

def get_soup_from_url_playwright(url, wait_for_selector='form', timeout=10000):
    """Load a page with Playwright and return BeautifulSoup of the rendered HTML."""
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()
        try:
            page.goto(url, timeout=timeout)
            try:
                page.wait_for_selector(wait_for_selector, timeout=5000)
            except Exception:
                pass  # If no form appears, just continue
            html = page.content()
        except Exception as e:
            print(f"Playwright failed to load {url}: {e}")
            html = ""
        finally:
            browser.close()
    return BeautifulSoup(html, 'html.parser') if html else None

def parse_args():
    parser = argparse.ArgumentParser(description='Website Test Case Generator')
    parser.add_argument('url', help='Website URL or GitHub Repo')
    parser.add_argument('--username', help='Username for login forms', default=None)
    parser.add_argument('--password', help='Password for login forms', default=None)
    return parser.parse_args()

def run_ddt_logins(url, login_excel='test_logins.xlsx', output_excel='test_cases_ddt.xlsx'):
    wb = openpyxl.load_workbook(login_excel)
    ws = wb.active
    results = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        username, password = row
        if PLAYWRIGHT_AVAILABLE:
            soup = get_soup_from_url_playwright(url)
        else:
            soup = get_soup_from_url(url)
        if not soup:
            results.append({'Username': username, 'Password': password, 'Type': '', 'Action': '', 'Element': '', 'Expected Result': '', 'Actual Result': 'Failed to load page', 'Notes': ''})
            continue
        test_cases = extract_elements(soup, url, username, password)
        for tc in test_cases:
            tc['Username'] = username
            tc['Password'] = password
            results.append(tc)
    # Write results to Excel with color coding
    from openpyxl.styles import PatternFill, Font
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    headers = ['Username', 'Password', 'Test Case ID', 'Type', 'Action', 'Element', 'Expected Result', 'Actual Result', 'Notes']
    ws_out.append(headers)
    # Define color fills
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    fill1 = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')
    fill2 = PatternFill(start_color='B8CCE4', end_color='B8CCE4', fill_type='solid')
    fill3 = PatternFill(start_color='95B3D7', end_color='95B3D7', fill_type='solid')
    fill4 = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    error_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    fills = [fill1, fill2, fill3, fill4]
    # Define fonts
    header_font = Font(bold=True, color='FFFFFF', size=14)
    data_font = Font(size=14)
    error_font = Font(size=14, color='FFFFFF', bold=True)
    # Apply header fill and font
    for col in range(1, len(headers) + 1):
        cell = ws_out.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
    # Data rows
    for idx, tc in enumerate(results, 1):
        row = [
            tc.get('Username', ''),
            tc.get('Password', ''),
            idx,
            tc.get('Type', ''),
            tc.get('Action', ''),
            tc.get('Element', ''),
            tc.get('Expected Result', ''),
            tc.get('Actual Result', ''),
            tc.get('Notes', '')
        ]
        ws_out.append(row)
        # Error highlighting
        actual_result = str(tc.get('Actual Result', '')).lower()
        is_error = (
            'broken' in actual_result or
            'not working' in actual_result or
            'failed' in actual_result or
            'error' in actual_result
        )
        fill = error_fill if is_error else fills[(idx - 1) % len(fills)]
        font = error_font if is_error else data_font
        for col in range(1, len(headers) + 1):
            cell = ws_out.cell(row=idx+1, column=col)
            cell.fill = fill
            cell.font = font
    wb_out.save(output_excel)
    print(f"DDT login test cases written to {output_excel}")

def main():
    args = parse_args()
    arg = args.url
    username = args.username
    password = args.password
    if arg.startswith('http') and 'github.com' in arg:
        analyze_github_repo(arg)
    elif arg.startswith('http'):
        if username == 'DDT' and password == 'DDT':
            run_ddt_logins(arg)
            return
        if PLAYWRIGHT_AVAILABLE:
            soup = get_soup_from_url_playwright(arg)
        else:
            soup = get_soup_from_url(arg)
        if not soup:
            print("Failed to analyze the website.")
            sys.exit(1)
        test_cases = extract_elements(soup, arg, username, password)
        write_to_excel(test_cases)
    else:
        print("Invalid argument. Please provide a website URL or GitHub repo URL.")
        sys.exit(1)

def get_soup_from_url(url):
    try:
        response = requests.get(url, timeout=10)
        response.raise_for_status()
        return BeautifulSoup(response.text, 'html.parser')
    except Exception as e:
        print(f"Failed to fetch {url}: {e}")
        return None

if __name__ == "__main__":
    main() 